import csv
import io
import os
import re
import urllib.request
import urllib.parse
import html as html_lib
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any

import fitz
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
try:
    import xlrd
except Exception:
    xlrd = None
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell, CoveredTableCell
from odf import teletype
from odf.namespaces import TABLENS

APP_VERSION = "4.0.0"
MAX_BYTES = 30 * 1024 * 1024
SUPPORTED = ["pdf", "xlsx", "xls", "xlsm", "ods", "csv", "tsv", "txt"]

app = FastAPI(title="OpsPilot Universal Parser", version=APP_VERSION)
origins = [
    "https://mcknight0521.github.io",
    "http://localhost:8000",
    "http://127.0.0.1:8000",
]
extra = os.getenv("OPSPILOT_FRONTEND_ORIGIN", "").strip()
if extra and extra not in origins:
    origins.append(extra)
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

DATE_DMY_RE = re.compile(r"\b(\d{2})/(\d{2})/(\d{4})\b")
DATE_YMD_RE = re.compile(r"\b(\d{4})[-/](\d{1,2})[-/](\d{1,2})\b")
SKU_RE = re.compile(r"(?m)^([0-9]{6})$")
REF_RE = re.compile(r"F\d{6,}")
REASON_RE = re.compile(r"\b([A-Z]-[A-Za-z][A-Za-z-]*)\b")
NUM_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?$")
TIME_RE = re.compile(r"^\d{2}:\d{2}:\d{2}$")

ALIASES = {
    "date": ["日期", "營運日期", "交易日期", "銷售日期", "調整日", "date", "sales date", "business date"],
    "sku": ["單品編號", "商品編號", "商品代碼", "品號", "料號", "sku", "item code", "product code"],
    "item": ["單品名稱", "商品名稱", "品名", "品項", "名稱", "item name", "product name"],
    "qty": ["銷售量", "銷售件數", "數量", "銷量", "qty", "quantity", "sales qty"],
    "sales": ["銷售額(含稅)", "含稅營業額", "營業額", "銷售額", "銷售淨額", "sales", "revenue", "amount", "net revenue"],
    "salesExTax": ["銷售額(未稅)", "未稅營業額", "未稅銷售額", "sales ex tax", "net sales"],
    "tax": ["稅額", "tax", "vat"],
    "waste": ["報廢", "報廢金額", "報廢損失", "損耗", "損耗金額", "waste", "scrap"],
    "clearance": ["降價出清金額", "出清金額", "折價出清", "出清", "clearance", "markdown amount"],
    "clearanceQty": ["降價出清數量", "出清數量", "clearance qty", "markdown qty"],
    "traffic": ["來客數", "入店人數", "客數", "traffic", "visitors", "customers"],
    "laborHours": ["工時", "人員工時", "總工時", "labor hours", "work hours"],
    "department": ["部門", "部門名稱", "課別", "department"],
    "category": ["分類", "分類別", "商品分類", "category"],
    "store": ["門市", "店別", "店名", "store"],
}



def clean(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v if v is not None else "").replace("\u00a0", " ")).strip()


def norm_header(v: Any) -> str:
    return re.sub(r"[\s　]+", "", clean(v)).lower()


def num(v: Any) -> float:
    if v is None or clean(v) == "":
        return 0.0
    s = clean(v).replace(",", "").replace("$", "").replace("％", "%")
    if s.endswith("%"):
        s = s[:-1]
    try:
        return float(s)
    except Exception:
        return 0.0


def to_iso(v: Any):
    if v is None or clean(v) == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = clean(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    # Conservative fallback for common date-time strings. Avoid heavyweight pandas
    # so the parser stays inside Render Free memory limits.
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def close(a, b, tol=.02):
    try:
        return abs(float(a) - float(b)) <= max(tol, abs(float(b)) * 1e-7)
    except Exception:
        return False


def vcheck(name, source, calc, tol=.02):
    if source is None:
        return {"metric": name, "status": "unverified", "sourceTotal": None, "recalculatedTotal": round(calc, 2), "difference": None}
    ok = close(calc, source, tol)
    return {
        "metric": name,
        "status": "passed" if ok else "failed",
        "sourceTotal": round(float(source), 2),
        "recalculatedTotal": round(float(calc), 2),
        "difference": round(float(calc) - float(source), 2),
    }


def summarize_checks(checks, **extra):
    failed = [c for c in checks if c.get("status") == "failed"]
    partial = [c for c in checks if c.get("status") in ("partial", "unverified")]
    status = "failed" if failed else ("partial" if partial else "passed")
    out = {
        "status": status,
        "checks": checks,
        "passed": sum(c.get("status") == "passed" for c in checks),
        "failed": len(failed),
        "unverified": sum(c.get("status") == "unverified" for c in checks),
        "warning": sum(c.get("status") == "partial" for c in checks),
    }
    out.update(extra)
    return out


def extract_period_text(text: str):
    m = re.search(r"調整日\s*(\d{2}/\d{2}/\d{4})\s*迄\s*(\d{2}/\d{2}/\d{4})", text, re.S)
    if m:
        return {"start": to_iso(m.group(1)), "end": to_iso(m.group(2))}
    dates = [m.group(0) for m in DATE_DMY_RE.finditer(text)]
    if len(dates) >= 2:
        # Printed date often appears first; use earliest and latest unique date when possible.
        iso = [to_iso(x) for x in dates]
        iso = sorted({x for x in iso if x})
        if len(iso) >= 2:
            return {"start": iso[0], "end": iso[-1]}
    return {"start": None, "end": None}


def cluster_blocks(blocks, rotation):
    axis = 0 if rotation in (90, 270) else 1
    items = [(float(b[axis]), b) for b in blocks if clean(b[4])]
    items.sort(key=lambda z: z[0])
    groups, centers = [], []
    for coord, b in items:
        idx = next((i for i, c in enumerate(centers) if abs(coord - c) <= 1.25), None)
        if idx is None:
            groups.append([b])
            centers.append(coord)
        else:
            groups[idx].append(b)
            n = len(groups[idx])
            centers[idx] = (centers[idx] * (n - 1) + coord) / n
    return groups


def parse_inventory_pdf(data: bytes, name: str):
    pdf = fitz.open(stream=data, filetype="pdf")
    try:
        full = "\n".join(p.get_text("text", sort=False) for p in pdf)
        if not re.search(r"庫存\s*調整\s*單", full):
            raise ValueError("unsupported")
        period = extract_period_text(full)
        rows, reasons, raw = [], set(), 0
        for page_no, page in enumerate(pdf, start=1):
            blocks = page.get_text("blocks", sort=False)
            page_text = "\n".join(str(b[4]) for b in blocks)
            rm = REASON_RE.search(page_text)
            reason = rm.group(1) if rm else ""
            if reason:
                reasons.add(reason)
            for group in cluster_blocks(blocks, page.rotation):
                fullg = "\n".join(str(b[4]).strip() for b in sorted(group, key=lambda b: (b[1], b[0])) if clean(b[4]))
                sm = SKU_RE.search(fullg)
                if not sm:
                    continue
                raw += 1
                sku = sm.group(1)
                ref = REF_RE.search(fullg)
                dates = [m.group(0) for m in DATE_DMY_RE.finditer(fullg)]
                if not ref or not dates:
                    continue
                item = ""
                for b in group:
                    lines = [x.strip() for x in str(b[4]).splitlines() if x.strip()]
                    for i, t in enumerate(lines[:-1]):
                        if TIME_RE.fullmatch(t):
                            item = clean(lines[i + 1])
                            break
                    if item:
                        break
                sku_lines = []
                for b in group:
                    lines = [x.strip() for x in str(b[4]).splitlines() if x.strip()]
                    if sku in lines:
                        sku_lines = lines
                        break
                try:
                    si = sku_lines.index(sku)
                except Exception:
                    continue
                detail = sku_lines[si - 1] if si > 0 and re.fullmatch(r"\d{3}", sku_lines[si - 1]) else ""
                vals = [num(x) for x in sku_lines[: max(0, si - 1)] if NUM_RE.fullmatch(x)]
                if len(vals) < 3:
                    continue
                amount, stock_before, qty = vals[-3:]
                unit = stock_after = diff = None
                flag = en = ""
                for b in group:
                    lines = [x.strip() for x in str(b[4]).splitlines() if x.strip()]
                    if "20" not in lines or not any(x in ("Y", "N") for x in lines):
                        continue
                    vs = [num(x) for x in lines if NUM_RE.fullmatch(x)]
                    if len(vs) >= 4 and abs(vs[-1] - 20) < 1e-9:
                        unit, diff, stock_after = vs[-2], vs[-3], vs[-4]
                        flag = next((x for x in lines if x in ("Y", "N")), "")
                        en = next((clean(x) for x in lines if not NUM_RE.fullmatch(x) and x not in ("Y", "N") and x != "20"), "")
                        break
                if unit is None:
                    continue
                waste = round(-amount, 2) if (reason.startswith("E-") or reason.startswith("B-")) and amount < 0 else None
                rows.append({
                    "date": to_iso(dates[0]), "sku": sku, "item": item or en or sku, "englishName": en or None,
                    "qty": None, "sales": None, "waste": waste, "clearance": None,
                    "periodStart": period["start"], "periodEnd": period["end"],
                    "adjustmentReason": reason or "未辨識", "adjustmentQty": qty,
                    "adjustmentAmount": round(amount, 2), "unitCost": unit,
                    "stockBefore": stock_before, "stockAfter": stock_after, "detailCode": detail,
                    "clearanceFlag": flag, "diffPct": diff, "ref": ref.group(0), "postDate": to_iso(dates[-1]),
                    "sourceType": "api-pdf-inventory-adjustment", "sourceFile": name, "sourcePage": page_no,
                })
        if not rows:
            raise ValueError("unparsed")
        math_ok = sum(1 for r in rows if close(r["adjustmentAmount"], r["adjustmentQty"] * r["unitCost"], .05))
        checks = [{
            "metric": "adjustment_amount_math",
            "status": "passed" if math_ok == len(rows) else "partial",
            "matchedRows": math_ok,
            "totalRows": len(rows),
        }]
        validation = summarize_checks(checks, rawTransactionCount=raw, parsedTransactions=len(rows), coverage=round(len(rows) / raw, 4) if raw else 1)
        return {
            "ok": True, "documentType": "inventory_adjustment", "confidence": 1.0 if validation["status"] == "passed" else .95,
            "validation": validation, "rows": rows,
            "meta": {"parser": "backend-pymupdf-inventory-v3", "reportType": "庫存調整單", "sourceType": "api-pdf", "sourceFile": name,
                     "pages": len(pdf), "parsedTransactions": len(rows), "rawTransactionCount": raw,
                     "coverage": round(len(rows) / raw, 4) if raw else 1, "adjustmentReasons": sorted(reasons),
                     "periodStart": period["start"], "periodEnd": period["end"]},
        }
    finally:
        pdf.close()


def parse_sales_pdf(data: bytes, name: str):
    pdf = fitz.open(stream=data, filetype="pdf")
    try:
        full = "\n".join(p.get_text("text", sort=False) for p in pdf)
        if "每日銷售報表 - 依單品" not in full:
            raise ValueError("unsupported")
        period = extract_period_text(full)
        # The last page grand total is a stable audit anchor for this report family.
        mtotal = re.search(r"總計\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+(-?[\d,]+)\s+([\d,.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)", full, re.S)
        source_total = None
        if mtotal:
            vals = [num(x) for x in mtotal.groups()]
            # Extracted order: ex-tax, tax, incl-tax, stock, qty, margin, benefit, immediate qty, immediate amt, clearance qty, clearance amt
            source_total = {"salesExTax": vals[0], "tax": vals[1], "sales": vals[2], "qty": vals[4], "clearanceQty": vals[9], "clearance": vals[10]}

        rows_raw = []
        for page_no, page in enumerate(pdf, start=1):
            lines = [x.strip() for x in page.get_text("text", sort=False).splitlines() if x.strip()]
            i = 0
            while i < len(lines):
                if not re.fullmatch(r"\d{6}", lines[i]):
                    i += 1
                    continue
                sku = lines[i]
                # Do not parse if this SKU is immediately part of a header-like block; require N/P marker before next SKU.
                j = i + 1
                while j < len(lines) and not re.fullmatch(r"\d{6}", lines[j]) and lines[j] not in ("小分類總計", "總計"):
                    j += 1
                block = lines[i + 1:j]
                np_idx = next((k for k, x in enumerate(block) if x in ("N", "P")), None)
                if np_idx is None or np_idx < 9 or len(block) < np_idx + 5:
                    i = max(j, i + 1)
                    continue
                pre = block[:np_idx]
                post = block[np_idx + 1:]
                # Expect final numeric scaffold before N/P: qty, ex-tax, tax, sales, GM, benefit, stock, dept, detail
                nums_pre = [x for x in pre if NUM_RE.fullmatch(x)]
                nums_post = [x for x in post if NUM_RE.fullmatch(x)]
                if len(nums_pre) < 9 or len(nums_post) < 4:
                    i = max(j, i + 1)
                    continue
                n = [num(x) for x in nums_pre[-9:]]
                p = [num(x) for x in nums_post[:4]]
                qty, ex_tax, tax, sales, gm, benefit, stock, dept, detail = n
                immediate_qty, immediate_amt, clearance_qty, clearance_amt = p
                # Name text is everything before the numeric scaffold, with Chinese/English separated heuristically.
                first_num_pos = next((k for k, x in enumerate(pre) if NUM_RE.fullmatch(x)), len(pre))
                names = pre[:first_num_pos]
                item = names[0] if names else sku
                if len(names) > 1 and any('\u4e00' <= ch <= '\u9fff' for ch in names[1]):
                    # wrapped Chinese product name
                    item = "".join(x for x in names if any('\u4e00' <= ch <= '\u9fff' for ch in x)) or item
                english = " ".join(x for x in names if not any('\u4e00' <= ch <= '\u9fff' for ch in x)) or None
                rows_raw.append({
                    "date": None, "sku": sku, "item": clean(item), "englishName": clean(english) if english else None, "np": block[np_idx],
                    "qty": qty, "sales": sales, "salesExTax": ex_tax, "tax": tax, "waste": None,
                    "clearance": clearance_amt, "clearanceQty": clearance_qty, "immediateClearanceQty": immediate_qty,
                    "immediateClearanceAmount": immediate_amt, "grossMarginPct": gm, "commercialBenefitPct": benefit,
                    "stock": stock, "periodStart": period["start"], "periodEnd": period["end"],
                    "sourceType": "api-pdf-daily-sales", "sourceFile": name, "sourcePage": page_no,
                })
                i = max(j, i + 1)
        if not rows_raw:
            raise ValueError("unparsed")
        result = finalize_daily_sales(rows_raw, name, "api-pdf", None, source_total, excluded_summary_rows=None)
        result["meta"]["parser"] = "backend-pymupdf-daily-sales-v3"
        result["meta"]["pages"] = len(pdf)
        return result
    finally:
        pdf.close()


def split_item(v):
    parts = [x.strip() for x in str(v or "").splitlines() if x.strip()]
    return (parts[0] if parts else "", " ".join(parts[1:]) if len(parts) > 1 else None)


def find_header(table):
    required = {norm_header(x) for x in ["單品編號", "單品名稱", "銷售量", "銷售額(含稅)"]}
    for i, row in enumerate(table[:150]):
        vals = {norm_header(x) for x in row if clean(x)}
        if required.issubset(vals):
            return i
    return None


def header_map(header):
    return {norm_header(h): i for i, h in enumerate(header) if clean(h)}


def hidx(idx, name):
    return idx.get(norm_header(name))


def finalize_daily_sales(detail, name, source_type, sheet, source_totals=None, excluded_summary_rows=0):
    merged = {}
    for r in detail:
        k = r["sku"]
        if k not in merged:
            merged[k] = {**r, "np": None, "qty": 0.0, "sales": 0.0, "salesExTax": 0.0, "tax": 0.0, "clearance": 0.0, "clearanceQty": 0.0}
        m = merged[k]
        for f in ("qty", "sales", "salesExTax", "tax", "clearance", "clearanceQty"):
            m[f] = (m.get(f) or 0) + (r.get(f) or 0)
    rows = list(merged.values())
    calc = {
        "qty": sum(r.get("qty") or 0 for r in rows),
        "sales": sum(r.get("sales") or 0 for r in rows),
        "salesExTax": sum(r.get("salesExTax") or 0 for r in rows),
        "tax": sum(r.get("tax") or 0 for r in rows),
        "clearance": sum(r.get("clearance") or 0 for r in rows),
        "clearanceQty": sum(r.get("clearanceQty") or 0 for r in rows),
    }
    checks = []
    for metric in ("sales", "qty", "clearance"):
        checks.append(vcheck(metric, source_totals.get(metric) if source_totals else None, calc[metric], 1.05 if metric == "sales" else .05))
    if source_totals and source_totals.get("salesExTax") is not None and source_totals.get("tax") is not None:
        checks.append(vcheck("sales_ex_tax", source_totals.get("salesExTax"), calc["salesExTax"], 1.05))
        checks.append(vcheck("tax", source_totals.get("tax"), calc["tax"], 1.05))
    tax_bad = sum(1 for r in rows if not close((r.get("salesExTax") or 0) + (r.get("tax") or 0), r.get("sales") or 0, 1.05))
    checks.append({"metric": "row_tax_math", "status": "passed" if tax_bad == 0 else "failed", "matchedRows": len(rows) - tax_bad, "totalRows": len(rows)})
    dup_removed = len(detail) - len(rows)
    validation = summarize_checks(
        checks,
        excludedSummaryRows=excluded_summary_rows,
        rawDetailRows=len(detail),
        mergedSkuRows=len(rows),
        duplicateNpRowsMerged=dup_removed,
    )
    validation["dataCompleteness"] = _completeness(rows, rows[0].get("periodStart") if rows else None, rows[0].get("periodEnd") if rows else None, "report_header")
    if validation["status"] == "failed":
        raise HTTPException(status_code=422, detail={"message": "報表驗算失敗，為避免錯誤數字，本次不匯入。", "validation": validation})
    return {
        "ok": True,
        "documentType": "daily_sales_by_item",
        "confidence": 1.0 if validation["status"] == "passed" else .90,
        "validation": validation,
        "totals": {k: round(v, 2) for k, v in calc.items()},
        "rows": rows,
        "meta": {
            "parser": "backend-daily-sales-v4", "reportType": "每日銷售報表 - 依單品", "sourceType": source_type,
            "sourceFile": name, "sheet": sheet, "periodStart": rows[0].get("periodStart") if rows else None,
            "periodEnd": rows[0].get("periodEnd") if rows else None, "rawDetailRows": len(detail), "parsedProducts": len(rows),
            "excludedSummaryRows": excluded_summary_rows,
        },
    }


def parse_daily_sales_table(table, name, source_type, sheet=None):
    hi = find_header(table)
    if hi is None:
        raise ValueError("unsupported")
    header = [clean(x) for x in table[hi]]
    idx = header_map(header)
    needed = ["單品編號", "單品名稱", "銷售量", "銷售額(含稅)", "降價出清金額"]
    if any(hidx(idx, x) is None for x in needed):
        raise ValueError("unsupported")
    start = end = None
    for row in table[:hi]:
        vals = [clean(x) for x in row]
        # Robustly collect dates in the report header; use earliest/latest.
        dates = [to_iso(x) for x in vals]
        dates = [x for x in dates if x]
        if dates:
            start = min([start] + dates) if start else min(dates)
            end = max([end] + dates) if end else max(dates)
    detail, summary_rows, source_totals = [], 0, None
    for row in table[hi + 1:]:
        vals = [clean(x) for x in row]
        if any(x in ("小分類總計", "總計") for x in vals):
            summary_rows += 1
            if "總計" in vals and "小分類總計" not in vals:
                source_totals = {
                    "qty": num(row[hidx(idx, "銷售量")]), "sales": num(row[hidx(idx, "銷售額(含稅)")]),
                    "salesExTax": num(row[hidx(idx, "銷售額(未稅)")]) if hidx(idx, "銷售額(未稅)") is not None else None,
                    "tax": num(row[hidx(idx, "稅額")]) if hidx(idx, "稅額") is not None else None,
                    "clearance": num(row[hidx(idx, "降價出清金額")]),
                    "clearanceQty": num(row[hidx(idx, "降價出清數量")]) if hidx(idx, "降價出清數量") is not None else None,
                }
            continue
        si = hidx(idx, "單品編號")
        sku = clean(row[si]) if si is not None and si < len(row) else ""
        if not re.fullmatch(r"\d{6}", sku):
            continue
        item, en = split_item(row[hidx(idx, "單品名稱")])
        np_i = hidx(idx, "N/P")
        detail.append({
            "date": None, "sku": sku, "item": item or sku, "englishName": en,
            "np": clean(row[np_i]) if np_i is not None else None,
            "qty": num(row[hidx(idx, "銷售量")]), "sales": num(row[hidx(idx, "銷售額(含稅)")]),
            "salesExTax": num(row[hidx(idx, "銷售額(未稅)")]) if hidx(idx, "銷售額(未稅)") is not None else 0,
            "tax": num(row[hidx(idx, "稅額")]) if hidx(idx, "稅額") is not None else 0,
            "waste": None, "clearance": num(row[hidx(idx, "降價出清金額")]),
            "clearanceQty": num(row[hidx(idx, "降價出清數量")]) if hidx(idx, "降價出清數量") is not None else 0,
            "periodStart": start, "periodEnd": end, "sourceType": source_type, "sourceFile": name,
        })
    if not detail:
        raise ValueError("unparsed")
    return finalize_daily_sales(detail, name, source_type, sheet, source_totals, summary_rows)


def _ods_cell_value(cell):
    text = clean(teletype.extractText(cell))
    if text:
        return text
    for attr in ("value", "datevalue", "timevalue", "booleanvalue"):
        try:
            v = cell.getAttribute(attr)
            if v not in (None, ""):
                return v
        except Exception:
            pass
    return None


def _ods_table_rows(table_node):
    rows = []
    # Cap repeated blank rows/columns defensively; malicious or highly formatted
    # ODS files can otherwise expand to huge in-memory matrices.
    MAX_REPEAT = 10000
    for row_node in table_node.getElementsByType(TableRow):
        try:
            rrep = int(row_node.getAttribute("numberrowsrepeated") or 1)
        except Exception:
            rrep = 1
        rrep = max(1, min(rrep, MAX_REPEAT))
        row = []
        for cell in row_node.childNodes:
            qname = getattr(cell, "qname", None)
            if qname not in ((TABLENS, "table-cell"), (TABLENS, "covered-table-cell")):
                continue
            try:
                crep = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            except Exception:
                crep = 1
            crep = max(1, min(crep, 512))
            value = _ods_cell_value(cell) if qname == (TABLENS, "table-cell") else None
            row.extend([value] * crep)
        # Trim trailing empty cells to keep memory bounded.
        while row and clean(row[-1]) == "":
            row.pop()
        for _ in range(rrep):
            rows.append(list(row))
    return rows


def load_spreadsheet_tables(data: bytes, ext: str):
    if ext in (".xlsx", ".xlsm"):
        # read_only=True is critical on small Render instances: openpyxl's normal
        # mode materializes styles/cells and can push the process over memory.
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        try:
            out = []
            for ws in wb.worksheets:
                # Some ERP exports incorrectly declare A1:A1 as the worksheet
                # dimension even though thousands of cells exist. In read-only
                # mode that would make openpyxl return only A1, so force a
                # dimension rescan when the declared range is suspicious.
                if (ws.max_row or 0) <= 1 and (ws.max_column or 0) <= 1:
                    try:
                        ws.reset_dimensions()
                    except Exception:
                        pass
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                out.append((ws.title, rows))
            return out
        finally:
            wb.close()
    if ext == ".xls":
        if xlrd is None:
            raise ValueError("XLS 解析套件未安裝")
        book = xlrd.open_workbook(file_contents=data, on_demand=True)
        out = []
        try:
            for sheet in book.sheets():
                rows = []
                for rx in range(sheet.nrows):
                    vals = []
                    for cx in range(sheet.ncols):
                        cell = sheet.cell(rx, cx)
                        v = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                v = xlrd.xldate_as_datetime(v, book.datemode)
                            except Exception:
                                pass
                        vals.append(v)
                    rows.append(vals)
                out.append((sheet.name, rows))
            return out
        finally:
            try:
                book.release_resources()
            except Exception:
                pass
    if ext == ".ods":
        doc = load_ods(io.BytesIO(data))
        out = []
        for tbl in doc.spreadsheet.getElementsByType(Table):
            name = tbl.getAttribute("name") or "Sheet"
            out.append((name, _ods_table_rows(tbl)))
        return out
    raise ValueError("unsupported")

def decode_text(data: bytes):
    for enc in ("utf-8-sig", "cp950", "big5", "utf-8", "utf-16", "latin1"):
        try:
            return data.decode(enc), enc
        except Exception:
            pass
    raise ValueError("encoding")


def text_to_table(data: bytes, ext: str):
    text, enc = decode_text(data)
    if ext == ".tsv":
        delim = "\t"
    elif ext == ".csv":
        sample = text[:20000]
        try:
            delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except Exception:
            delim = ","
    else:
        sample = text[:20000]
        try:
            delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except Exception:
            # TXT with fixed-ish spacing: convert 2+ spaces to tabs as a conservative fallback.
            if any("\t" in x for x in text.splitlines()[:20]):
                delim = "\t"
            else:
                text = "\n".join(re.sub(r" {2,}", "\t", ln.rstrip()) for ln in text.splitlines())
                delim = "\t"
    return list(csv.reader(io.StringIO(text), delimiter=delim)), enc, delim


def match_alias(header_value: Any):
    h = norm_header(header_value)
    if not h:
        return None
    for field, aliases in ALIASES.items():
        for a in aliases:
            if h == norm_header(a):
                return field
    return None


def find_generic_header(table):
    best = None
    for i, row in enumerate(table[:150]):
        mapping = {}
        for j, v in enumerate(row):
            f = match_alias(v)
            if f and f not in mapping:
                mapping[f] = j
        score = len(mapping)
        if score >= 2 and ("sales" in mapping or "waste" in mapping or "clearance" in mapping or "qty" in mapping):
            if best is None or score > best[0]:
                best = (score, i, mapping)
    return best


def _date_range(start, end):
    if not start or not end:
        return []
    try:
        a = datetime.fromisoformat(start).date(); b = datetime.fromisoformat(end).date()
    except Exception:
        return []
    if b < a or (b-a).days > 3660:
        return []
    return [(a + timedelta(days=i)).isoformat() for i in range((b-a).days+1)]


def _extract_period_from_tables(tables, filename=""):
    dates=[]
    text=[]
    for sheet, table in tables:
        for row in table[:80]:
            for v in row:
                if v is None: continue
                text.append(clean(v))
                iso=to_iso(v)
                if iso: dates.append(iso)
    joined=" ".join(text)
    # Explicit ranges like 2026/09/01～2026/09/30 or 01/08/2026 ... 30/08/2026
    for pat in [r'(20\d{2})[/-](\d{1,2})[/-](\d{1,2})\s*[～~至-]\s*(20\d{2})[/-](\d{1,2})[/-](\d{1,2})',
                r'(\d{1,2})/(\d{1,2})/(20\d{2}).{0,30}?(\d{1,2})/(\d{1,2})/(20\d{2})']:
        m=re.search(pat, joined)
        if m:
            g=list(map(int,m.groups()))
            try:
                if len(g)==6 and g[0]>=2000: return date(g[0],g[1],g[2]).isoformat(), date(g[3],g[4],g[5]).isoformat(), 'report_header'
                return date(g[2],g[1],g[0]).isoformat(), date(g[5],g[4],g[3]).isoformat(), 'report_header'
            except Exception: pass
    # Filename range, useful for exports that omit period metadata.
    m=re.search(r'(20\d{2})(\d{2})(\d{2})[-_~至]+(20\d{2})(\d{2})(\d{2})', filename)
    if m:
        g=list(map(int,m.groups()))
        try: return date(g[0],g[1],g[2]).isoformat(), date(g[3],g[4],g[5]).isoformat(), 'filename'
        except Exception: pass
    return (min(dates), max(dates), 'observed_dates') if dates else (None,None,None)


def _completeness(rows, period_start=None, period_end=None, period_source=None):
    actual=sorted({r.get('date') for r in rows if r.get('date')})
    if not actual:
        return {"status":"not_applicable","reason":"此報表沒有可安全辨識的逐日日期。"}
    start=period_start or actual[0]; end=period_end or actual[-1]
    expected=_date_range(start,end)
    missing=[d for d in expected if d not in set(actual)]
    extra=[d for d in actual if expected and d not in set(expected)]
    return {
        "status":"complete" if not missing and not extra else "incomplete",
        "periodStart":start,"periodEnd":end,"periodSource":period_source or "observed_dates",
        "expectedDays":len(expected),"actualDays":len(actual),"missingDays":missing,"extraDays":extra,
        "missingCount":len(missing),"note":"缺少日期不會自動視為 0 元，也不會自行推定停業原因。" if missing else None,
    }


def _daily_dimension_summary(rows, field):
    by={}
    for r in rows:
        d=r.get('date'); v=r.get(field)
        if not d or v is None: continue
        by.setdefault(d,[]).append(v)
    if not by: return None
    daily={}; ambiguous=[]; strategy={}
    for d, vals in by.items():
        uniq=[]
        for v in vals:
            if v not in uniq: uniq.append(v)
        if len(uniq)==1:
            daily[d]=uniq[0]; strategy[d]='deduplicated_daily_value'
        else:
            # Multiple different values at item grain cannot safely be called store-day traffic/labor.
            daily[d]=None; ambiguous.append(d); strategy[d]='ambiguous_item_level_values'
    return {"dailyValues":daily,"ambiguousDates":ambiguous,"safeDays":sum(v is not None for v in daily.values()),
            "status":"verified" if not ambiguous else "partial",
            "rule":"同日期各商品值完全相同時只計一次；不同時不擅自加總。"}

def parse_generic_table(table, name, source_type, sheet=None, global_period=None):
    found = find_generic_header(table)
    if not found:
        raise ValueError("unsupported")
    score, hi, mapping = found
    if "sku" not in mapping and "item" not in mapping:
        raise ValueError("unsupported")
    rows=[]; excluded=0
    for row in table[hi+1:]:
        vals=[clean(x) for x in row]
        if any(re.search(r"(^|\s)(小計|總計|subtotal|grand total)(\s|$)",x,re.I) for x in vals if x):
            excluded+=1; continue
        sku=clean(row[mapping["sku"]]) if "sku" in mapping and mapping["sku"]<len(row) else ""
        item=clean(row[mapping["item"]]) if "item" in mapping and mapping["item"]<len(row) else ""
        d=to_iso(row[mapping["date"]]) if "date" in mapping and mapping["date"]<len(row) else None
        # Require a real data identity; footer/note rows are ignored.
        if not sku and not item: continue
        if "date" in mapping and not d: continue
        def val(f): return num(row[mapping[f]]) if f in mapping and mapping[f] < len(row) else None
        rows.append({
            "date":d,"sku":sku or item,"item":item or sku,"englishName":None,
            "qty":val("qty"),"sales":val("sales"),"salesExTax":val("salesExTax"),"tax":val("tax"),
            "waste":val("waste"),"clearance":val("clearance"),"clearanceQty":val("clearanceQty"),
            "traffic":val("traffic"),"laborHours":val("laborHours"),
            "department":clean(row[mapping["department"]]) if "department" in mapping else None,
            "category":clean(row[mapping["category"]]) if "category" in mapping else None,
            "store":clean(row[mapping["store"]]) if "store" in mapping else None,
            "sourceType":source_type,"sourceFile":name,
        })
    if not rows: raise ValueError("unparsed")
    gp=global_period or (None,None,None); ps,pe,psrc=gp
    dates=sorted({r["date"] for r in rows if r.get("date")})
    ps=ps or (dates[0] if dates else None); pe=pe or (dates[-1] if dates else None)
    for r in rows: r["periodStart"]=ps; r["periodEnd"]=pe
    checks=[]
    if "salesExTax" in mapping and "tax" in mapping and "sales" in mapping:
        checked=[r for r in rows if r.get("sales") is not None]
        bad=sum(1 for r in checked if not close((r.get("salesExTax") or 0)+(r.get("tax") or 0),r.get("sales") or 0,1.05))
        checks.append({"metric":"row_tax_math","status":"passed" if bad==0 else "partial","matchedRows":len(checked)-bad,"totalRows":len(checked)})
    checks.append({"metric":"source_grand_total","status":"unverified","reason":"通用格式未找到可安全辨識的原始總計；核心欄位採重算並標示部分驗證。"})
    completeness=_completeness(rows,ps,pe,psrc)
    traffic=_daily_dimension_summary(rows,"traffic") if "traffic" in mapping else None
    labor=_daily_dimension_summary(rows,"laborHours") if "laborHours" in mapping else None
    confidence=min(.96,.52+score*.055)
    validation=summarize_checks(checks,mappedFields=sorted(mapping.keys()),excludedSummaryRows=excluded,parsedRows=len(rows))
    validation["dataCompleteness"]=completeness
    validation["dailyDimensions"]={"traffic":traffic,"laborHours":labor}
    totals={f:(round(sum(r.get(f) or 0 for r in rows),2) if f in mapping else None) for f in ("sales","qty","waste","clearance")}
    # Only expose traffic/labor total when daily values are unambiguous.
    for f,obj in (("traffic",traffic),("laborHours",labor)):
        totals[f]=round(sum(v for v in (obj or {}).get("dailyValues",{}).values() if v is not None),2) if obj and not obj.get("ambiguousDates") else None
    return {"ok":True,"documentType":"generic_operational_table","confidence":confidence,"validation":validation,"totals":totals,"rows":rows,
            "meta":{"parser":"backend-trust-engine-v4","reportType":"通用營運表格","sourceType":source_type,"sourceFile":name,"sheet":sheet,
                    "periodStart":ps,"periodEnd":pe,"periodSource":psrc,"mappedFields":sorted(mapping.keys()),"requiresReview":validation["status"]!="passed",
                    "trustEngine":"v4"}}

def parse_tabular(data: bytes, name: str, ext: str):
    if ext in (".xlsx", ".xlsm", ".xls", ".ods"):
        tables = load_spreadsheet_tables(data, ext)
        source_type = f"api-{ext[1:]}"
    else:
        table, enc, delim = text_to_table(data, ext)
        tables = [(None, table)]
        source_type = f"api-{ext[1:] or 'txt'}"
    global_period = _extract_period_from_tables(tables, name)
    # First pass: known report fingerprints. Second pass: generic mapping.
    errors = []
    for sheet, table in tables:
        try:
            return parse_daily_sales_table(table, name, source_type, sheet)
        except HTTPException:
            raise
        except Exception as e:
            errors.append(f"{sheet or 'text'}:known:{e}")
    candidates = []
    for sheet, table in tables:
        try:
            candidates.append(parse_generic_table(table, name, source_type, sheet, global_period))
        except Exception as e:
            errors.append(f"{sheet or 'text'}:generic:{e}")
    if candidates:
        return sorted(candidates, key=lambda x: x.get("confidence", 0), reverse=True)[0]
    raise ValueError("unsupported")


def parse_any(data: bytes, name: str):
    ext = Path(name).suffix.lower()
    if ext == ".pdf":
        # Try known PDF report families in deterministic order.
        for parser in (parse_inventory_pdf, parse_sales_pdf):
            try:
                return parser(data, name)
            except HTTPException:
                raise
            except ValueError as e:
                if str(e) != "unsupported":
                    raise
        raise HTTPException(status_code=422, detail="PDF 已讀取，但目前無法安全辨識此報表；本次資料未匯入。")
    if ext in (".xlsx", ".xls", ".xlsm", ".ods", ".csv", ".tsv", ".txt"):
        return parse_tabular(data, name, ext)
    raise HTTPException(status_code=415, detail=f"目前接受：{', '.join(SUPPORTED).upper()}")


@app.get("/")
def root():
    return {"ok": True, "service": "OpsPilot Universal Parser", "version": APP_VERSION, "docs": "/docs", "health": "/health"}




COUNTY_NAMES = [
    "基隆市","臺北市","新北市","桃園市","新竹市","新竹縣","苗栗縣","臺中市","彰化縣","南投縣",
    "雲林縣","嘉義市","嘉義縣","臺南市","高雄市","屏東縣","宜蘭縣","花蓮縣","臺東縣","澎湖縣","金門縣","連江縣"
]

def _http_text(url: str, timeout: int = 20) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 OpsPilot/3.2"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
        enc = r.headers.get_content_charset() or "utf-8"
        try:
            return raw.decode(enc, errors="replace")
        except Exception:
            return raw.decode("utf-8", errors="replace")

def _strip_html(v: str) -> str:
    v = re.sub(r"<script[\s\S]*?</script>", " ", v or "", flags=re.I)
    v = re.sub(r"<style[\s\S]*?</style>", " ", v, flags=re.I)
    v = re.sub(r"<[^>]+>", " ", v)
    return re.sub(r"\s+", " ", html_lib.unescape(v)).strip()

def _roc_to_iso(y: str, m: str, d: str) -> str:
    return f"{int(y)+1911:04d}-{int(m):02d}-{int(d):02d}"

def _extract_history_links(page_html: str, base: str) -> list[dict]:
    out, seen = [], set()
    for m in re.finditer(r'<a[^>]+href=["\']([^"\']*information\?[^"\']*pid=\d+[^"\']*)["\'][^>]*>([\s\S]*?)</a>', page_html or '', re.I):
        href, label = m.group(1), _strip_html(m.group(2))
        around = _strip_html((page_html or '')[max(0,m.start()-260):m.end()+260])
        dm = re.search(r'(\d{2,3})年\s*(\d{1,2})月\s*(\d{1,2})日', label + ' ' + around)
        if not dm:
            continue
        url = urllib.parse.urljoin(base, html_lib.unescape(href))
        key = (url, dm.group(0))
        if key in seen:
            continue
        seen.add(key)
        out.append({"date": _roc_to_iso(*dm.groups()), "url": url, "title": label or around[:120]})
    return out

def _extract_nds_links(detail_html: str, base: str) -> list[str]:
    out=[]
    for m in re.finditer(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>([\s\S]*?)</a>', detail_html or '', re.I):
        href, label = html_lib.unescape(m.group(1)), _strip_html(m.group(2))
        if re.search(r'nds(?:E)?\.html|停止上班|停止辦公|停止上課', href+' '+label, re.I) and not re.search(r'ndsE\.html', href, re.I):
            u=urllib.parse.urljoin(base, href)
            if u not in out: out.append(u)
    return out

def _parse_region_closure(nds_html: str, region: str) -> dict:
    plain = _strip_html(nds_html)
    idx = plain.find(region)
    if idx < 0:
        return {"found": False, "closure": False, "detail": ""}
    tail = plain[idx+len(region):]
    stops=[]
    for county in COUNTY_NAMES:
        j=tail.find(county)
        if j>=0: stops.append(j)
    chunk = tail[:min(stops) if stops else 700].strip()
    closure = bool(re.search(r'停止上班|停止上課|停止辦公|停班|停課', chunk)) and not bool(re.search(r'照常上班.{0,40}照常上課|正常上班.{0,40}正常上課', chunk))
    scope = "partial"
    if closure and re.match(r'^[：: ]*(?:今天|明天)?停止上班[、,， ]*(?:今天|明天)?停止上課', chunk):
        scope = "countywide"
    return {"found": True, "closure": closure, "detail": chunk[:500], "scope": scope}

def _dgpa_closures(start: str, end: str, region: str) -> list[dict]:
    base='https://www.dgpa.gov.tw'
    targets=[]
    start_dt=datetime.strptime(start,'%Y-%m-%d').date()
    end_dt=datetime.strptime(end,'%Y-%m-%d').date()
    for page in range(1,51):
        url=f'{base}/informationlist?page={page}&uid=374'
        try:
            text=_http_text(url)
        except Exception:
            if page>3: break
            continue
        items=_extract_history_links(text,url)
        if not items:
            if page>3: break
            continue
        page_dates=[]
        for it in items:
            try: d=datetime.strptime(it['date'],'%Y-%m-%d').date()
            except Exception: continue
            page_dates.append(d)
            if start_dt <= d <= end_dt:
                targets.append(it)
        if page_dates and min(page_dates) < start_dt:
            break
    uniq={x['url']:x for x in targets}.values()
    out=[]
    for it in uniq:
        try:
            detail=_http_text(it['url'])
        except Exception:
            continue
        nds_links=_extract_nds_links(detail,it['url'])
        parsed=None
        for u in nds_links:
            try:
                parsed=_parse_region_closure(_http_text(u),region)
            except Exception:
                continue
            if parsed.get('found'): break
        if parsed and parsed.get('closure'):
            out.append({
                "type":"closure", "date":it['date'],
                "name":"全縣停班停課" if parsed.get('scope')=='countywide' else "部分地區停班停課",
                "detail":parsed.get('detail',''), "scope":parsed.get('scope','partial'),
                "source":"行政院人事行政總處"
            })
    return sorted(out,key=lambda x:x['date'])

@app.get("/history-events")
def history_events(start: str, end: str, region: str = "屏東縣"):
    try:
        datetime.strptime(start, "%Y-%m-%d")
        datetime.strptime(end, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="start/end must be YYYY-MM-DD")
    if region not in COUNTY_NAMES:
        raise HTTPException(status_code=422, detail="unsupported region")
    try:
        closures=_dgpa_closures(start,end,region)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"DGPA history lookup failed: {e}")
    return {"ok":True,"region":region,"start":start,"end":end,"events":closures,"sources":["行政院人事行政總處"]}

@app.get("/health")
def health():
    return {"ok": True, "service": "opspilot-parser", "version": APP_VERSION, "engine": "Trust Engine v4", "formats": SUPPORTED, "endpoint": "/parse-file", "multiFileEndpoint": "/verify-files", "historyEndpoint": "/history-events"}


@app.post("/verify-files")
async def verify_files(files: list[UploadFile] = File(...)):
    if not files or len(files) > 8:
        raise HTTPException(status_code=400, detail="請上傳 1～8 個檔案。")
    results=[]
    for f in files:
        data=await f.read()
        if len(data)>MAX_BYTES:
            results.append({"file":f.filename,"ok":False,"error":"file_too_large"}); continue
        try: results.append({"file":f.filename,**parse_any(data,f.filename or "upload")})
        except Exception as e: results.append({"file":f.filename,"ok":False,"error":str(e)})
    comparable=[x for x in results if x.get("ok") and x.get("totals")]
    pairs=[]
    for i in range(len(comparable)):
        for j in range(i+1,len(comparable)):
            a,b=comparable[i],comparable[j]; metrics={}
            for m in ("sales","qty","waste","clearance"):
                av=a.get("totals",{}).get(m); bv=b.get("totals",{}).get(m)
                if av is not None and bv is not None:
                    metrics[m]={"a":av,"b":bv,"difference":round(av-bv,2),"matched":close(av,bv,1.05 if m in ("sales","waste","clearance") else .05)}
            if metrics: pairs.append({"fileA":a.get("file"),"fileB":b.get("file"),"metrics":metrics,"matched":all(v["matched"] for v in metrics.values())})
    return {"ok":True,"version":APP_VERSION,"results":results,"crossFileValidation":{"pairs":pairs,"status":"passed" if pairs and all(p["matched"] for p in pairs) else ("partial" if pairs else "not_applicable")}}

@app.post("/parse-file")
async def parse_file(file: UploadFile = File(...)):
    name = file.filename or "upload"
    data = await file.read()
    if not data:
        raise HTTPException(400, "空白檔案")
    if len(data) > MAX_BYTES:
        raise HTTPException(413, "檔案超過 30MB")
    try:
        return parse_any(data, name)
    except HTTPException:
        raise
    except ValueError as e:
        if str(e) == "unsupported":
            raise HTTPException(422, "目前無法安全辨識此報表；可改用欄位確認流程建立新模板，本次不會把未確認數字寫入正式分析。")
        raise HTTPException(422, f"檔案已讀取，但驗證或解析失敗：{e}")
    except Exception as e:
        raise HTTPException(400, f"檔案無法開啟：{e}")


# Backward compatibility for the current website while it migrates to /parse-file.
@app.post("/parse-pdf")
async def parse_pdf(file: UploadFile = File(...)):
    name = file.filename or "upload.pdf"
    data = await file.read()
    if not name.lower().endswith(".pdf"):
        raise HTTPException(415, "只接受 PDF 檔案")
    try:
        return parse_any(data, name)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(422, f"PDF 解析失敗：{e}")
